"""
Admin Panel for DMTT Application Bot
A Flask web application to view and manage applications
"""

from flask import Flask, render_template, send_file, jsonify, request
import csv
import os
from datetime import datetime
from dotenv import load_dotenv
from telegram import Bot
import asyncio
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

app = Flask(__name__)
app.config['SECRET_KEY'] = os.urandom(24)

# Configuration
CSV_FILE = "applications.csv"
BOT_TOKEN = os.getenv("BOT_TOKEN")
DOWNLOAD_FOLDER = "downloaded_pdfs"

# Create download folder if it doesn't exist
os.makedirs(DOWNLOAD_FOLDER, exist_ok=True)

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def read_applications():
    """Read all applications from CSV file"""
    applications = []
    
    if not os.path.exists(CSV_FILE):
        return applications
    
    try:
        with open(CSV_FILE, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for idx, row in enumerate(reader, 1):
                row['id'] = idx
                applications.append(row)
        
        # Sort by date (newest first)
        applications.reverse()
        
    except Exception as e:
        print(f"Error reading CSV: {e}")
    
    return applications

def get_application_stats():
    """Get statistics about applications"""
    applications = read_applications()
    
    total = len(applications)
    today = datetime.now().strftime('%Y-%m-%d')
    today_count = sum(1 for app in applications if app['Sana'].startswith(today))
    
    return {
        'total': total,
        'today': today_count,
        'last_updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }

async def download_pdf_async(file_id, filename):
    """Download PDF from Telegram asynchronously"""
    try:
        bot = Bot(token=BOT_TOKEN)
        file = await bot.get_file(file_id)
        
        # Create safe filename
        safe_filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}"
        filepath = os.path.join(DOWNLOAD_FOLDER, safe_filename)
        
        await file.download_to_drive(filepath)
        return filepath
    except Exception as e:
        print(f"Error downloading PDF: {e}")
        return None

def download_pdf(file_id, filename):
    """Synchronous wrapper for PDF download"""
    return asyncio.run(download_pdf_async(file_id, filename))

def create_excel_export():
    """Create a formatted Excel file from applications data"""
    applications = read_applications()
    
    if not applications:
        return None
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Arizalar"
    
    # Define headers
    headers = ['#', 'Sana', 'Ism', 'Telefon', 'Username', 'Chat ID', 'File ID', 'Fayl nomi']
    
    # Header styling
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Write data
    for row_num, app in enumerate(applications, 2):
        # Reverse order to show newest first
        data = [
            app.get('id', ''),
            app.get('Sana', ''),
            app.get('Ism', ''),
            app.get('Telefon', ''),
            app.get('Username', ''),
            app.get('Chat ID', ''),
            app.get('File ID', ''),
            app.get('Fayl nomi', '')
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            
            # Alternate row colors
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="F8F9FF", end_color="F8F9FF", fill_type="solid")
    
    # Auto-adjust column widths
    column_widths = {
        'A': 8,   # #
        'B': 20,  # Sana
        'C': 25,  # Ism
        'D': 18,  # Telefon
        'E': 20,  # Username
        'F': 15,  # Chat ID
        'G': 40,  # File ID
        'H': 30   # Fayl nomi
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save to file
    filename = f"DMTT_Arizalar_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(DOWNLOAD_FOLDER, filename)
    wb.save(filepath)
    
    return filepath

# ============================================================================
# ROUTES
# ============================================================================

@app.route('/')
def index():
    """Main dashboard page"""
    applications = read_applications()
    stats = get_application_stats()
    return render_template('index.html', applications=applications, stats=stats)

@app.route('/api/applications')
def api_applications():
    """API endpoint to get all applications as JSON"""
    applications = read_applications()
    return jsonify(applications)

@app.route('/api/stats')
def api_stats():
    """API endpoint to get statistics"""
    stats = get_application_stats()
    return jsonify(stats)

@app.route('/download/<int:app_id>')
def download(app_id):
    """Download PDF for a specific application"""
    applications = read_applications()
    
    if app_id < 1 or app_id > len(applications):
        return "Application not found", 404
    
    app = applications[app_id - 1]
    file_id = app.get('File ID')
    filename = app.get('Fayl nomi', 'document.pdf')
    
    if not file_id:
        return "File ID not found", 404
    
    # Download PDF from Telegram
    filepath = download_pdf(file_id, filename)
    
    if filepath and os.path.exists(filepath):
        return send_file(filepath, as_attachment=True, download_name=filename)
    else:
        return "Error downloading file", 500

@app.route('/export')
def export_excel():
    """Export all applications as formatted Excel file"""
    filepath = create_excel_export()
    
    if filepath and os.path.exists(filepath):
        filename = os.path.basename(filepath)
        return send_file(filepath, as_attachment=True, download_name=filename)
    else:
        return "No applications found", 404

@app.route('/search')
def search():
    """Search applications"""
    query = request.args.get('q', '').lower()
    
    if not query:
        return jsonify([])
    
    applications = read_applications()
    results = []
    
    for app in applications:
        # Search in name, phone, and username
        if (query in app.get('Ism', '').lower() or 
            query in app.get('Telefon', '').lower() or 
            query in app.get('Username', '').lower()):
            results.append(app)
    
    return jsonify(results)

# ============================================================================
# MAIN
# ============================================================================

if __name__ == '__main__':
    print("🌐 Admin Panel ishga tushmoqda...")
    print("📊 Dashboard: http://localhost:5000")
    print("🛑 To'xtatish uchun Ctrl+C bosing")
    print()
    app.run(debug=True, host='0.0.0.0', port=5000)
